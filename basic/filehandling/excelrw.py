#  using openpyxl for excel operation in python
from openpyxl import load_workbook

# giving excel path
f = load_workbook(filename="D:\\pythonproject\\basic\\data\\excel.xlsx")
# fetching all the sheets in excel
sh=f.sheetnames
# going to retrieve first sheet data
sheet= f[sh[0]]
c =sheet.cell(row=1, column=1)
username = c.value
c =sheet.cell(row=2, column=2)
password = c.value
print("Username :" +username)
print("Password :" +password)

# going to retrieve first sheet data
sheet= f[sh[1]]
# printing sheet name
print(sheet.title)

# empty list
item = []
Quantity = []

# number of rows in excel
i = sheet.max_row
for j in range(1, i):
    c = sheet.cell(row=j+1, column=1)
    item.insert(j, c.value)
    c = sheet.cell(row=j+1, column=2)
    Quantity.insert(j, c.value)

print(item)
print(Quantity)
f.close()